import pandas as pd
from selenium.common.exceptions import StaleElementReferenceException, NoSuchElementException, WebDriverException, TimeoutException, UnexpectedAlertPresentException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium import webdriver
from datetime import datetime
import time
from selenium.webdriver.support.ui import Select
import threading
import numpy as np
import sys
import traceback
import eel
import logging # Adicionado para consistência, já que é usado nas funções de template.
import subprocess
import pymysql # USANDO PyMySQL
import pymysql.cursors
import win32gui
import win32con
import os
import ctypes
import queue
import glob
import openpyxl
from openpyxl.styles import Font, Alignment # Importa Font e Alignment
import shutil # Importado para clonar arquivos
from urllib.parse import urlparse # Importado para normalizar URLs
from math import ceil

# Configuração de logging (opcional, mas bom para depuração)
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Função para obter o caminho correto do driver quando compilado
def get_driver_path():
    if getattr(sys, 'frozen', False):
        # Rodando como um executável PyInstaller
        return os.path.join(sys._MEIPASS, 'msedgedriver.exe')
    # Rodando como script normal
    return 'msedgedriver.exe'


def normalize_url(url_string):
    """
    Normaliza uma string de URL para comparação, removendo parâmetros de consulta
    e garantindo um formato consistente.
    """
    if not url_string:
        return ""
    try:
        parsed_url = urlparse(url_string)
        # Reconstrói o URL sem a query string e o fragmento
        normalized_path = parsed_url.scheme + "://" + parsed_url.netloc + parsed_url.path
        # Remove barra final se houver
        if normalized_path.endswith('/'):
            normalized_path = normalized_path.rstrip('/')
        return normalized_path.lower() # Converte para minúsculas para comparação case-insensitive
    except Exception as e:
        logging.warning(f"Erro ao normalizar URL '{url_string}': {e}. Retornando string original.")
        return str(url_string).lower() # Retorna a string original em minúsculas como fallback
    
@eel.expose   
def open_results_foldervectora():
    """
    Abre a pasta de resultados da Macro Consulta Geral no explorador de arquivos.
    A pasta é Desktop/Macro JGL/Macro Consulta Geral.
    """
    try:
        home_dir = os.path.expanduser('~')
        results_path = os.path.join(home_dir, 'Desktop', 'Macro JGL', 'Macro Deliberação')

        if not os.path.exists(results_path):
            os.makedirs(results_path, exist_ok=True)

        if sys.platform == "win32" or os.name == 'nt':
            os.startfile(results_path)
        elif sys.platform == "darwin":
            subprocess.run(['open', results_path], check=True)
        else:
            subprocess.run(['xdg-open', results_path], check=True)

        return {"status": "success", "message": f"Pasta '{results_path}' aberta com sucesso."}
    except Exception as e:
        error_message = f"Erro ao abrir a pasta de resultados: {e}"
        print(error_message)
        return {"status": "error", "message": error_message}

def carregar_arquivo(caminho_arquivo):
    """
    Carrega o arquivo Excel original e retorna o primeiro valor da coluna 'instalação'.
    Este arquivo é o template.
    """
    try:
        # ATENÇÃO: Verifique se o caminho do arquivo está correto e se o nome está correto.
        # No seu código original, o nome era "excel_da_deliberaçãoteste.xlsx".
        # No seu último código, você mudou para "excel_da_deliberação.xlsx".
        # Mantenho o que você forneceu no último código.
        df = pd.read_excel(caminho_arquivo)
        if df.empty:
            print("Arquivo template está vazio. Fim do processamento.")
            return None, None # Retorna None para df e primeiro_valor
        
        nomes_possiveis = ['instalação', 'INSTALACAO', 'instalacao', 'INSTALAÇÃO', 'Instalação']
        nome_coluna = next((col for col in nomes_possiveis if col in df.columns), None)
        
        if nome_coluna:
            return df, nome_coluna # Retorna o DataFrame completo e o nome da coluna de instalação
        else:
            print("Nenhuma coluna de instalação encontrada no arquivo template.")
            return df, None # Retorna o DataFrame, mas sem a coluna de instalação
    except FileNotFoundError:
        print("Arquivo template não encontrado.")
        return None, None
    except Exception as e:
        print(f"Erro ao carregar o arquivo template: {str(e)}")
        return None, None

def get_column_index_by_name(ws, column_name):
    """
    Retorna o índice da coluna (base 1) dado o nome do cabeçalho.
    Retorna -1 se a coluna não for encontrada.
    """
    for col_index in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col_index).value == column_name:
            return col_index
    return -1

def armazenar_dados_proc_final(numero_serie, link_numero_serie, idx_linha_template, df_resultados, ws_resultados, wb_resultados, output_file_path, nome_col_equipamento, nome_col_status):
    """
    Adiciona/Atualiza os dados de equipamento e status no DataFrame de resultados
    e aplica a formatação no worksheet do openpyxl, usando o índice da linha do template
    para posicionamento.

    Args:
        numero_serie (str): O número de série encontrado.
        link_numero_serie (str): O link associado ao número de série.
        idx_linha_template (int): O índice da linha no DataFrame original (template)
                                   que corresponde à instalação processada.
        df_resultados (pd.DataFrame): O DataFrame que está sendo construído para o arquivo de saída.
        ws_resultados (openpyxl.worksheet.worksheet.Worksheet): O worksheet do openpyxl do arquivo de saída.
        wb_resultados (openpyxl.workbook.workbook.Workbook): O workbook do openpyxl do arquivo de saída.
        output_file_path (str): O caminho completo do arquivo de saída.
        nome_col_equipamento (str): Nome da coluna de equipamento (padrão 'EQUIPAMENTO').
        nome_col_status (str): Nome da coluna de status (padrão 'STATUS').

    Returns:
        bool: True se o item foi considerado "Atualizado" (link/equipamento mudou), False se "Inalterado" (novo ou igual).
    """
    is_updated_status = False # Define se o status será "Atualizado" ou "Inalterado"
    try:
        linha_para_usar_df = idx_linha_template

        # Garante que o DataFrame de resultados tenha a linha para ser atualizada
        if linha_para_usar_df >= len(df_resultados):
            num_linhas_faltantes = linha_para_usar_df - len(df_resultados) + 1
            empty_rows = pd.DataFrame([[np.nan] * len(df_resultados.columns)] * num_linhas_faltantes, columns=df_resultados.columns)
            df_resultados = pd.concat([df_resultados, empty_rows], ignore_index=True)
        
        # Encontra o índice da coluna de EQUIPAMENTO no openpyxl (base 1)
        col_idx_equipamento_ws = get_column_index_by_name(ws_resultados, nome_col_equipamento)
        if col_idx_equipamento_ws == -1:
            logging.error(f"Erro crítico: Coluna '{nome_col_equipamento}' não encontrada no worksheet de resultados. Isso não deveria acontecer após a inicialização na macro.")
            return False # Impede que continue com erros
        
        # Garante que a linha existe no worksheet para evitar IndexError ao acessar a célula
        row_openpyxl = linha_para_usar_df + 2 # +1 para índice base 1, +1 para linha de cabeçalho
        if row_openpyxl > ws_resultados.max_row:
            for r in range(ws_resultados.max_row + 1, row_openpyxl + 1):
                for c in range(1, ws_resultados.max_column + 1):
                    ws_resultados.cell(row=r, column=c, value="")
        
        current_equipamento_ws_cell = ws_resultados.cell(row=row_openpyxl, column=col_idx_equipamento_ws)
        
        # Obter o valor atual do equipamento do DataFrame (confiável)
        current_equipamento_df_value = df_resultados.at[linha_para_usar_df, nome_col_equipamento]
        
        # NORMALIZAÇÃO DOS LINKS PARA COMPARAÇÃO
        normalized_new_link = normalize_url(link_numero_serie)
        normalized_current_link = normalize_url(current_equipamento_ws_cell.hyperlink.target) if current_equipamento_ws_cell.hyperlink else ""

        # --- CORREÇÃO: Se número de série está vazio, marca como Inconsistencia ---
        if not numero_serie or str(numero_serie).strip() == '' or str(numero_serie).lower() == 'none':
            status_text = 'Inconsistencia'
            is_updated_status = False
            # Atualiza DataFrame
            df_resultados.at[linha_para_usar_df, nome_col_equipamento] = ''
            df_resultados.at[linha_para_usar_df, nome_col_status] = status_text
            # Atualiza worksheet
            cell_equipamento = ws_resultados.cell(row=row_openpyxl, column=col_idx_equipamento_ws)
            cell_equipamento.value = ''
            cell_equipamento.hyperlink = None
            cell_equipamento.font = Font(color="FF0000")
            cell_equipamento.alignment = Alignment(horizontal='center', vertical='center')
            col_idx_status_ws = get_column_index_by_name(ws_resultados, nome_col_status)
            if col_idx_status_ws != -1:
                cell_status = ws_resultados.cell(row=row_openpyxl, column=col_idx_status_ws, value=status_text)
                cell_status.font = Font(color="FF0000")
                cell_status.alignment = Alignment(horizontal='center', vertical='center')
            wb_resultados.save(output_file_path)
            logging.info(f"Linha {row_openpyxl}: Inconsistencia - número de série vazio.")
            return False
        # --- FIM DA CORREÇÃO ---

        # Lógica de "Alterado" vs "Inalterado"
        if pd.isna(current_equipamento_df_value) or str(current_equipamento_df_value).strip() == '':
            # Equipamento novo para esta linha (célula estava vazia no DataFrame)
            status_text = 'Inalterado'
            is_updated_status = False
        else:
            # Equipamento já existia na linha. Verifica se o número de série ou o link mudou.
            if str(current_equipamento_df_value) != str(numero_serie):
                # O número de série mudou
                status_text = 'Alterado'
                is_updated_status = True
            else:
                # O número de série é o mesmo. Verifica se o link mudou (comparando links normalizados)
                if normalized_current_link == normalized_new_link:
                    # Mesmo número de série e mesmo link (normalizado)
                    status_text = 'Inalterado'
                    is_updated_status = False
                else:
                    # Mesmo número de série, mas link diferente (após normalização)
                    status_text = 'Alterado'
                    is_updated_status = True

        # Atualiza os valores no DataFrame de resultados
        df_resultados.at[linha_para_usar_df, nome_col_equipamento] = str(numero_serie)
        df_resultados.at[linha_para_usar_df, nome_col_status] = status_text

        # Atualiza a célula de equipamento no worksheet
        cell_equipamento = ws_resultados.cell(row=row_openpyxl, column=col_idx_equipamento_ws)
        center_alignment = Alignment(horizontal='center', vertical='center')
        cell_equipamento.alignment = center_alignment
        # --- CORREÇÃO: só coloca hyperlink se for realmente válido ---
        link_invalido = False
        if link_numero_serie:
            if "?inc=" in link_numero_serie or link_numero_serie.strip() == '' or link_numero_serie == 'None':
                link_invalido = True
        if is_updated_status:
            cell_equipamento.value = str(numero_serie) if numero_serie else ""
            if link_numero_serie and link_numero_serie.strip() and link_numero_serie != str(numero_serie) and not link_invalido:
                cell_equipamento.hyperlink = link_numero_serie
                cell_equipamento.font = Font(color="0000FF", underline="single")
            else:
                cell_equipamento.hyperlink = None
                cell_equipamento.font = Font(color="87CEEB")  # Azul céu claro
        else:
            cell_equipamento.value = str(numero_serie) if numero_serie else ""
            cell_equipamento.hyperlink = None
            cell_equipamento.font = Font(color="00B050")  # Verde claro
        # Atualiza a cor para a coluna EQUIPAMENTO
        if is_updated_status:
            light_blue_font = Font(color="87CEEB")  # Azul céu claro (SkyBlue)
            cell_equipamento.font = light_blue_font
        else:
            green_font = Font(color="00B050")  # Verde claro
            cell_equipamento.font = green_font
        # --- FIM DA CORREÇÃO ---
        # Atualiza a célula de status no worksheet
        col_idx_status_ws = get_column_index_by_name(ws_resultados, nome_col_status)
        if col_idx_status_ws != -1:
            cell_status = ws_resultados.cell(row=row_openpyxl, column=col_idx_status_ws, value=status_text)
            cell_status.alignment = center_alignment
            if is_updated_status:
                cell_status.font = light_blue_font
            else:
                cell_status.font = green_font
        else:
            logging.warning(f"Atenção: Coluna '{nome_col_status}' não encontrada no worksheet para atualizar o status. Isso não deveria acontecer.")

        # SALVA O WORKBOOK A CADA ITERAÇÃO
        wb_resultados.save(output_file_path)
        logging.info(f"Dados e formatação aplicados para {numero_serie} na linha {row_openpyxl}. Arquivo salvo.")
        return is_updated_status
        
    except Exception as e:
        print(f"Erro em armazenar_dados_proc_final para {numero_serie}: {str(e)}")
        traceback.print_exc()
        return False

def login():
    global driver, wait
    prefs = {"profile.default_content_setting_values.notifications": 2 }
    options = Options()
    
    # Certifique-se de que o caminho para o msedgedriver.exe está correto ou no PATH do sistema
    # Exemplo: driver = webdriver.Edge(executable_path='C:/caminho/para/msedgedriver.exe', options=options)
    service = Service(executable_path=get_driver_path())
    driver = webdriver.ChromiumEdge(service=service, options=options)
    wait = WebDriverWait(driver, 10)
    
    driver.get('http://10.7.41.190/vectorasys/')
    
    l_costumer = 'sabesp.mn'
    l_login ='lcordeiro'
    l_senha = 'eficien.sabesp'
 
    driver.maximize_window()

    try:
        time.sleep(2)
        costumer_field = driver.find_element(By.XPATH, "/html/body/div/form/table/tbody/tr[1]/td[2]/input")
        username_field = driver.find_element(By.XPATH, "/html/body/div/form/table/tbody/tr[2]/td[2]/input")
        password_field = driver.find_element(By.XPATH, "/html/body/div/form/table/tbody/tr[3]/td[2]/input[1]")
        login_button = driver.find_element(By.XPATH, "/html/body/div/form/table/tbody/tr[4]/td/button")
        
        costumer_field.send_keys(l_costumer)
        username_field.send_keys(l_login)
        password_field.send_keys(l_senha)
        login_button.click()
        time.sleep(2)

        # REMOVE a chamada do macro() daqui!
        # O login não deve abrir navegador nem chamar macro()
    except Exception as e:
         print(f"Erro ao interagir com a janela de autenticação: {str(e)}")
         traceback.print_exc()

def worker_processar_instalacoes(q, df_template, nome_col_instalacao, output_file_path, nome_col_equipamento_final, nome_col_status_final, lock, total_processar, contador_global, erros_global):
    options = Options()
    options.add_argument("--incognito")
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.page_load_strategy = 'normal'
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-extensions")
    options.add_argument("--guest")
    options.add_experimental_option("prefs", {
        'download.prompt_for_download': False,
        'download.directory_upgrade': True,
        'safeBrowse.enabled': True,
        "profile.default_content_setting_values.notifications": 2
    })
    service = Service(executable_path=get_driver_path())
    driver = webdriver.ChromiumEdge(service=service, options=options)
    wait = WebDriverWait(driver, 10)
    driver.get('http://10.7.41.190/vectorasys/')
    l_costumer = 'sabesp.mn'
    l_login ='lcordeiro'
    l_senha = 'eficien.sabesp'
    driver.maximize_window()
    tempo_inicial = time.time()  # Marca o início do processamento
    try:
        time.sleep(2)
        costumer_field = driver.find_element(By.XPATH, "/html/body/div/form/table/tbody/tr[1]/td[2]/input")
        username_field = driver.find_element(By.XPATH, "/html/body/div/form/table/tbody/tr[2]/td[2]/input")
        password_field = driver.find_element(By.XPATH, "/html/body/div/form/table/tbody/tr[3]/td[2]/input[1]")
        login_button = driver.find_element(By.XPATH, "/html/body/div/form/table/tbody/tr[4]/td/button")
        costumer_field.send_keys(l_costumer)
        username_field.send_keys(l_login)
        password_field.send_keys(l_senha)
        login_button.click()
        time.sleep(2)
        while True:
            try:
                idx_df_template, valor_instalacao = q.get_nowait()
            except queue.Empty:
                break
            ciclo_inicio = time.time()  # Marca o início do ciclo
            try:

                aba_scorpion = WebDriverWait(driver, 60).until(
                    EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[3]/div[2]/div/ul/li[2]/a"))
                )
                aba_scorpion.click()
                
                input_scorpion = WebDriverWait(driver, 60).until(
                    EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[3]/div[2]/div/div[2]/div/div[1]/div[2]/input"))
                )
         
                input_scorpion.clear()
                input_scorpion.send_keys(str(valor_instalacao))
                time.sleep(2)
                numero_serie_element = WebDriverWait(driver, 60).until(
                    EC.presence_of_element_located((By.XPATH, "/html/body/div[1]/div[3]/div[2]/div/div[2]/div/table/tbody/tr/td[1]/a"))
                )
                
                numero_serie = numero_serie_element.text
                link_numero_serie = numero_serie_element.get_attribute("href")
                time.sleep(2)
                with lock:
                    wb_resultados = openpyxl.load_workbook(output_file_path)
                    ws_resultados = wb_resultados.active
                    df_resultados = pd.read_excel(output_file_path)
                    is_updated_status = armazenar_dados_proc_final(
                        numero_serie,
                        link_numero_serie,
                        idx_df_template,
                        df_resultados,
                        ws_resultados,
                        wb_resultados,
                        output_file_path,
                        nome_col_equipamento_final,
                        nome_col_status_final
                    )
                # Defina status_text corretamente para passar ao eel
                if is_updated_status:
                    status_text = 'Alterado'
                else:
                    status_text = 'Inalterado'
                with contador_global.get_lock():
                    contador_global.value += 1
                    processados = contador_global.value
                restantes = total_processar - processados
                tempo_decorrido = time.time() - tempo_inicial
                tempo_medio = tempo_decorrido / processados if processados > 0 else 0
                tempo_estimado_restante = tempo_medio * restantes
                cor_status = "#87CEEB" if status_text == "Alterado" else "#00B050"
                atualizar_status_macro_front(
                    valor_instalacao,
                    status_text,
                    processados,
                    erros_global.value,
                    total_processar,
                    tempo_estimado_restante,
                    cor_status
                )
                atualizar_porcentagem_front(processados, total_processar)
            except Exception as e:
                with erros_global.get_lock():
                    erros_global.value += 1
                    erros_atual = erros_global.value
                with contador_global.get_lock():
                    contador_global.value += 1
                    processados = contador_global.value
                # Atualiza Excel com cor vermelha e status 'INCONSISTENCIA'
                with lock:
                    wb_resultados = openpyxl.load_workbook(output_file_path)
                    ws_resultados = wb_resultados.active
                    df_resultados = pd.read_excel(output_file_path)
                    row_openpyxl = idx_df_template + 2
                    col_idx_equipamento_ws = get_column_index_by_name(ws_resultados, nome_col_equipamento_final)
                    col_idx_status_ws = get_column_index_by_name(ws_resultados, nome_col_status_final)
                    if col_idx_equipamento_ws != -1:
                        cell_equipamento = ws_resultados.cell(row=row_openpyxl, column=col_idx_equipamento_ws)
                        cell_equipamento.value = "Inconsistencia"
                        cell_equipamento.font = Font(color="FF0000")
                        cell_equipamento.alignment = Alignment(horizontal='center', vertical='center')
                        cell_equipamento.hyperlink = None
                    if col_idx_status_ws != -1:
                        cell_status = ws_resultados.cell(row=row_openpyxl, column=col_idx_status_ws)
                        cell_status.value = "Inconsistencia"
                        cell_status.font = Font(color="FF0000")
                        cell_status.alignment = Alignment(horizontal='center', vertical='center')
                    wb_resultados.save(output_file_path)
                cor_status = "#FF0000"
                total_concluidos = processados
                restantes = total_processar - total_concluidos
                tempo_decorrido = time.time() - tempo_inicial
                tempo_medio = tempo_decorrido / total_concluidos if total_concluidos > 0 else 0
                tempo_estimado_restante = tempo_medio * restantes
                atualizar_status_macro_front(
                    valor_instalacao,
                    "Inconsistencia",
                    total_concluidos,
                    erros_atual,
                    total_processar,
                    tempo_estimado_restante,
                    cor_status
                )
                atualizar_porcentagem_front(total_concluidos, total_processar)
                print(f"[Thread {threading.current_thread().name}] ERRO ao processar a instalação {valor_instalacao}: {str(e)} | Processados: {contador_global.value}/{total_processar} | Restantes: {total_processar - contador_global.value} | Erros: {erros_global.value}")
                traceback.print_exc()
                continue  # Pula o incremento de processados!
            finally:
                q.task_done()
    finally:
        driver.quit()
        print(f"[Thread {threading.current_thread().name}] finalizou e navegador foi fechado.")

def macro(caminho_arquivo, costumer, login, senha, tipo_arquivo, identificador_usuario):
    global driver, wait
    try:
        # 1. Carregar o arquivo template
        df_template, nome_col_instalacao = carregar_arquivo(caminho_arquivo)
        if df_template is None:
            print("Não foi possível carregar o arquivo template ou ele está vazio. Encerrando.")
            return
        if nome_col_instalacao is None:
            print("Coluna 'INSTALAÇÃO' não encontrada no arquivo template. Encerrando.")
            return

        # 2. Definir caminhos para o arquivo de resultados
        home_dir = os.path.expanduser('~')
        output_dir = os.path.join(home_dir, 'Desktop', 'Macro JGL', 'Macro Deliberação')
        os.makedirs(output_dir, exist_ok=True)
        
        now = datetime.now()
        data_formatada = now.strftime("%d_%m_%Y")
        output_file_name = f'ResultadoDELIBERAÇÃO-{data_formatada}.xlsx'
        output_file_path = os.path.join(output_dir, output_file_name)

        wb_resultados = None 
        ws_resultados = None
        df_resultados = None

        # 3. Clonar o arquivo enviado para o arquivo de resultados (NÃO salvar o arquivo enviado na pasta, só o resultado)
        import shutil
        shutil.copy(caminho_arquivo, output_file_path)
        wb_resultados = openpyxl.load_workbook(output_file_path)
        ws_resultados = wb_resultados.active

        # 4. Normalizar nomes de colunas, posicionar 'STATUS' e aplicar negrito nos cabeçalhos
        nome_col_equipamento_final = 'EQUIPAMENTO'
        nome_col_status_final = 'STATUS'
        bold_font = Font(bold=True) # Fonte em negrito para cabeçalhos
        center_alignment_headers = Alignment(horizontal='center', vertical='center') # Alinhamento para cabeçalhos

        # --- Acha e normaliza a coluna EQUIPAMENTO ---
        current_equip_col_idx = get_column_index_by_name(ws_resultados, nome_col_equipamento_final)
        if current_equip_col_idx == -1: # Se não encontrou 'EQUIPAMENTO'
            # Tenta encontrar com nomes possíveis (case-insensitive)
            found_equip_col_name = None
            for col_idx in range(1, ws_resultados.max_column + 1):
                cell_value = ws_resultados.cell(row=1, column=col_idx).value
                if cell_value and str(cell_value).lower() in ['equipamento', 'equipamento', 'equipamento']:
                    found_equip_col_name = cell_value
                    current_equip_col_idx = col_idx
                    break
            
            if found_equip_col_name: # Se encontrou com outro nome, renomeia
                if found_equip_col_name != nome_col_equipamento_final:
                    ws_resultados.cell(row=1, column=current_equip_col_idx, value=nome_col_equipamento_final)
            else: # Se não encontrou nenhuma coluna de equipamento, adiciona no final por enquanto
                ws_resultados.cell(row=1, column=ws_resultados.max_column + 1, value=nome_col_equipamento_final)
                current_equip_col_idx = ws_resultados.max_column # Pega o novo índice
        
        # Aplica negrito e centraliza o cabeçalho da coluna EQUIPAMENTO
        equip_header_cell = ws_resultados.cell(row=1, column=current_equip_col_idx)
        equip_header_cell.font = bold_font
        equip_header_cell.alignment = center_alignment_headers


        # --- Acha e normaliza a coluna STATUS e posiciona ao lado de EQUIPAMENTO ---
        # A ordem deve ser EQUIPMENT, STATUS, então o índice do STATUS deve ser current_equip_col_idx + 1
        desired_status_col_idx = current_equip_col_idx + 1
        current_status_col_idx = get_column_index_by_name(ws_resultados, nome_col_status_final)

        if current_status_col_idx == -1: # Se a coluna STATUS não existe (nome exato)
            # Tenta encontrar com nomes possíveis (case-insensitive)
            found_status_col_name = None
            for col_idx in range(1, ws_resultados.max_column + 1):
                cell_value = ws_resultados.cell(row=1, column=col_idx).value
                if cell_value and str(cell_value).lower() in ['status', 'status', 'status']:
                    found_status_col_name = cell_value
                    current_status_col_idx = col_idx # Achou a coluna existente
                    break
            
            if found_status_col_name: # Se encontrou com outro nome, renomeia e move se necessário
                if found_status_col_name != nome_col_status_final:
                    ws_resultados.cell(row=1, column=current_status_col_idx, value=nome_col_status_final)
                
                # Se a coluna STATUS já existe, mas não está na posição desejada, move-a
                if current_status_col_idx != desired_status_col_idx:
                    # 1. Copiar dados da coluna STATUS (para todas as linhas)
                    status_col_data = []
                    for row_idx in range(1, ws_resultados.max_row + 1):
                        status_col_data.append(ws_resultados.cell(row=row_idx, column=current_status_col_idx).value)
                    
                    # 2. Deletar a coluna STATUS existente
                    ws_resultados.delete_cols(current_status_col_idx)
                    
                    # 3. Inserir uma nova coluna na posição desejada
                    ws_resultados.insert_cols(desired_status_col_idx)
                    
                    # 4. Colar os dados de volta na nova coluna
                    for row_idx, value in enumerate(status_col_data, start=1):
                        ws_resultados.cell(row=row_idx, column=desired_status_col_idx, value=value)
                    
                    current_status_col_idx = desired_status_col_idx # Atualiza o índice após a movimentação

            else: # Se não encontrou nenhuma coluna de status (nem exata, nem similar), insere nova
                ws_resultados.insert_cols(desired_status_col_idx)
                ws_resultados.cell(row=1, column=desired_status_col_idx, value=nome_col_status_final)
                current_status_col_idx = desired_status_col_idx # Atualiza o índice da coluna STATUS
        
        # Aplica negrito e centraliza o cabeçalho da coluna STATUS
        status_header_cell = ws_resultados.cell(row=1, column=current_status_col_idx)
        status_header_cell.font = bold_font
        status_header_cell.alignment = center_alignment_headers

        # Salva o workbook com os cabeçalhos normalizados/adicionados/posicionados
        wb_resultados.save(output_file_path)
        
        # Recarrega o df_resultados após as alterações de cabeçalho no openpyxl
        # Isso garante que o DataFrame em memória reflita as colunas como elas estão no arquivo.
        df_resultados = pd.read_excel(output_file_path)

        indices_valores = list(df_template[nome_col_instalacao].dropna().items())
        total_processar = len(indices_valores)
        print(f"Total de valores para processar: {total_processar}")
        tempo_inicial = datetime.now()
        print("Começo da operação: ", tempo_inicial.strftime("%H:%M -- %d/%m"))

        max_workers = 3  # Ajuste conforme sua máquina
        lock = threading.Lock()
        import multiprocessing
        contador_global = multiprocessing.Value('i', 0)
        erros_global = multiprocessing.Value('i', 0)
        q = queue.Queue()
        for item in indices_valores:
            if str(item[1]).strip() == '':
                continue
            q.put(item)
        threads = []
        for i in range(max_workers):
            t = threading.Thread(target=worker_processar_instalacoes, args=(q, df_template, nome_col_instalacao, output_file_path, nome_col_equipamento_final, nome_col_status_final, lock, total_processar, contador_global, erros_global))
            t.start()
            threads.append(t)
        q.join()
        for t in threads:
            t.join()
        print(f"\nResumo final: Processados: {contador_global.value}/{total_processar} | Erros: {erros_global.value}")
        print("----------------------------------")
        print("Operação realizada com sucesso")
        print("----------------------------------")
        print("Começo da operação: ", tempo_inicial.strftime("%H:%M -- %d/%m"))
        tempo_final = datetime.now()
        print("Termino da operação: ", tempo_final.strftime("%H:%M -- %d/%m"))
        print(f"Total de INSTALAÇÃO processadas ({total_processar})")
        tempo_total = tempo_final - tempo_inicial
        total_segundos = int(tempo_total.total_seconds())
        total_minutos = int(total_segundos // 60)
        total_horas = int(total_minutos // 60)
        if total_segundos < 60:
            print(f"Tempo total da operação: {total_segundos} segundos")
        else:
            total_minutos = total_segundos // 60
            total_segundos_restantes = total_segundos % 60
            if total_minutos < 60:
                print(f"Tempo total da operação: {total_minutos} minutos e {total_segundos_restantes} segundos")
            else:
                total_horas = total_minutos // 60
                total_minutos_restantes = total_minutos % 60
                print(f"Tempo total da operação: {total_horas} horas e {total_minutos_restantes} minutos")
        
        # Chama o resumo final na interface via eel
        try:
            # Formata o tempo total para string legível
            if total_horas > 0:
                tempo_total_legivel = f"{total_horas}h {total_minutos_restantes}m {total_segundos_restantes}s"
            elif total_minutos > 0:
                tempo_total_legivel = f"{total_minutos}m {total_segundos_restantes}s"
            else:
                tempo_total_legivel = f"{total_segundos}s"
            eel.mostrarResumoFinalDeliberacao({
                "inicio": tempo_inicial.strftime("%d/%m/%Y às %H:%M:%S"),
                "termino": tempo_final.strftime("%d/%m/%Y às %H:%M:%S"),
                "processados": contador_global.value,
                "erros": erros_global.value,
                "tempoTotal": tempo_total_legivel
            })
        except Exception as e:
            logging.error(f"Erro ao chamar mostrarResumoFinalDeliberacao via eel: {e}")
    except Exception as e:
        print(f"Ocorreu um erro na macro principal: {str(e)}")
        traceback.print_exc()

class WaitTimeoutError(Exception):
    pass

@eel.expose
def iniciar_macro_deliberacao(costumer, login, senha, conteudo_base64, nome_arquivo, tipo_arquivo, identificador_usuario):
    """
    Função exposta para ser chamada via Eel, inicia a macro de deliberação recebendo costumer, login, senha, conteúdo do arquivo, nome do arquivo, tipo do arquivo e identificador do usuário.
    """
    import logging
    import base64
    import os
    import tempfile
    logging.info(f"Iniciando macro de deliberação para usuário '{identificador_usuario}' com costumer '{costumer}', login '{login}' e arquivo '{nome_arquivo}'")
    try:
        # Salva o arquivo Excel enviado pelo usuário em um arquivo temporário
        with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(nome_arquivo)[-1]) as temp_file:
            temp_file.write(base64.b64decode(conteudo_base64.split(',')[-1]))
            caminho_arquivo_temp = temp_file.name
        print(f"Costumer: {costumer}, Login: {login}, Senha: {senha}, Caminho arquivo TEMP: {caminho_arquivo_temp}, Tipo: {tipo_arquivo}, Identificador: {identificador_usuario}")
        # Passe o arquivo temporário para a macro
        macro(caminho_arquivo_temp, costumer, login, senha, tipo_arquivo, identificador_usuario)
        # Após a execução, remova o arquivo temporário
        os.remove(caminho_arquivo_temp)
        return {"status": "sucesso", "message": "Macro de deliberação iniciada."}
    except Exception as e:
        logging.error(f"Erro ao iniciar macro de deliberação: {e}")
        return {"status": "erro", "message": str(e)}

def formatar_tempo_legivel(segundos):
    if segundos is None:
        return "Calculando..."
    segundos = int(segundos)
    if segundos < 0:
        return "Calculando..."
    horas = segundos // 3600
    minutos = (segundos % 3600) // 60
    segundos_restantes = segundos % 60
    partes = []
    if horas > 0:
        partes.append(f"{horas}h")
    if minutos > 0 or (horas > 0 and segundos_restantes > 0):
        partes.append(f"{minutos}m")
    if segundos_restantes > 0 or (not partes and segundos == 0):
        partes.append(f"{segundos_restantes}s")
    if not partes:
        return "0s"
    return ' '.join(partes)

# Função para atualizar status e tempo estimado juntos

@eel.expose
def atualizar_status_macro_front(processando, status, quantidade, erros, total, tempo_estimado_segundos=None, cor_status=None):
    tempo_legivel = formatar_tempo_legivel(tempo_estimado_segundos) if tempo_estimado_segundos is not None else None
    eel.atualizar_status_macro(
        str(processando),
        str(status),
        int(quantidade),
        int(erros),
        int(total),
        tempo_legivel,
        cor_status
    )

def atualizar_porcentagem_front(processados, total):
    if not total or total == 0:
        porcentagem = 0
    else:
        porcentagem = int((processados / total) * 100)
    eel.atualizar_porcentagem_concluida(f"{porcentagem}%")