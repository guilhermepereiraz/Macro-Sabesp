import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles.colors import BLUE
import re
import os
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException, ElementClickInterceptedException, StaleElementReferenceException
from selenium.webdriver.common.keys import Keys

from colorama import Fore, Style, init

# Inicializar colorama para habilitar cores no terminal
init(autoreset=True)

# --- Configurações da Planilha ---
caminho_planilha = r"C:\Users\lcordeiro.eficien.SBSP\Desktop\LCS\excel da deliberação.xlsx"
nome_aba = "deliberação"
coluna_instalacao = "INSTALAÇÃO"      # Contém o nome completo da VRP/PC para busca
coluna_equipamento = "Equipamento"    # Será atualizada com o NÚMERO DO EQUIPAMENTO (do title ou texto) e conterá a URL como hiperlink

# --- Configurações do Selenium ---
base_url_vectora = "http://10.7.41.190/vectorasys/"
caminho_chromedriver = r"C:\Users\lcordeiro.eficien.SBSP\Desktop\LCS\Drivers\chromedriver-win64\chromedriver.exe"
timeout_selenium_geral = 60 # Tempo limite maior para operações gerais
timeout_pequena_espera = 30 # Tempo limite para esperas mais curtas

# --- Seletores HTML ---
SELETOR_ABA_SCORPION = (By.LINK_TEXT, "Scorpion 35PF")
SELETOR_SEARCH_BOX_SCORPICO_ESPECIFICO = (By.XPATH, "//div[@id='prd2_filter']/input[@type='text']")
SELETOR_TABELA_RESULTS = (By.ID, "prd2_wrapper") # Usando o ID da div wrapper

# Seletores para Login
SELETOR_CUSTOMER_FIELD = (By.ID, "customer")
SELETOR_USERNAME_FIELD = (By.ID, "user")
SELETOR_PASSWORD_FIELD = (By.ID, "passwd")
SELETOR_LOGIN_BUTTON = (By.XPATH, "//button[text()=' Login']")

# --- SUAS CREDENCIAIS ---
CUSTOMER_VALUE = "sabesp.mn"
USERNAME_VALUE = "lcordeiro"
PASSWORD_VALUE = "eficien.sabesp"

# --- FUNÇÕES ---
def get_seletor_link_do_numero_fixo(numero_fixo_para_clicar):
    # ATENÇÃO: Busca pelo TEXTO VISÍVEL do link.
    # Usando LINK_TEXT por ser mais direto. Se houver problemas com espaços ou caracteres especiais,
    # considere a alternativa com XPATH e normalize-space().
    return (By.LINK_TEXT, numero_fixo_para_clicar)
    # Alternativa mais robusta se By.LINK_TEXT falhar (ex: texto com espaços extras):
    # return (By.XPATH, f"//div[@id='prd2_wrapper']//a[normalize-space()='{numero_fixo_do_pc_title}']")

def realizar_login(driver, customer, user, password, base_url, timeout):
    print(Fore.BLUE + "\n--- Tentando fazer login no VectoraSYS ---" + Style.RESET_ALL)
    driver.get(base_url)
    try:
        WebDriverWait(driver, timeout).until(EC.presence_of_element_located(SELETOR_CUSTOMER_FIELD))
        driver.find_element(*SELETOR_CUSTOMER_FIELD).send_keys(customer)
        driver.find_element(*SELETOR_USERNAME_FIELD).send_keys(user)
        driver.find_element(*SELETOR_PASSWORD_FIELD).send_keys(password)
        WebDriverWait(driver, timeout).until(EC.element_to_be_clickable(SELETOR_LOGIN_BUTTON)).click()
        print(Fore.GREEN + "     Botão 'Login' clicado. Aguardando redirecionamento..." + Style.RESET_ALL)

        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located(SELETOR_ABA_SCORPION)
        )
        print(Fore.GREEN + "     Login bem-sucedido." + Style.RESET_ALL)
        return True
    except TimeoutException as e:
        print(Fore.RED + f"     ERRO FATAL: Timeout excedido ao tentar fazer login. Verifique as credenciais, a URL, ou a conectividade. Detalhes: {e}" + Style.RESET_ALL)
        return False
    except NoSuchElementException as e:
        print(Fore.RED + f"     ERRO FATAL: Elemento de login não encontrado. Verifique os seletores ou se a página de login mudou. Detalhes: {e}" + Style.RESET_ALL)
        return False
    except WebDriverException as e:
        print(Fore.RED + f"     ERRO FATAL: Erro do WebDriver (ex: navegador fechou inesperadamente, desconexão). Detalhes: {e}" + Style.RESET_ALL)
        return False
    except Exception as e:
        print(Fore.RED + f"     ERRO inesperado durante o login: {e.__class__.__name__}: {e}" + Style.RESET_ALL)
        return False

def verificar_e_relogar(driver, customer, user, password, base_url, timeout):
    # Verifica se a URL atual é a página de login
    if driver.current_url == base_url or "login" in driver.current_url:
        print(Fore.YELLOW + "Sessão expirada ou redirecionado para o login. Tentando re-login..." + Style.RESET_ALL)
        if not realizar_login(driver, customer, user, password, base_url, timeout):
            raise Exception("Falha ao re-logar. Encerrando script.")
    else:
        try:
            # Tenta encontrar um elemento que só aparece após o login (ex: aba Scorpion)
            WebDriverWait(driver, timeout_pequena_espera).until(
                EC.presence_of_element_located(SELETOR_ABA_SCORPION)
            )
        except (TimeoutException, NoSuchElementException, StaleElementReferenceException):
            print(Fore.YELLOW + "Não foi possível confirmar a sessão ativa. Tentando re-login por segurança..." + Style.RESET_ALL)
            if not realizar_login(driver, customer, user, password, base_url, timeout):
                raise Exception("Falha ao re-logar. Encerrando script.")

# --- INÍCIO DO PROCESSO PRINCIPAL ---
try:
    print(Fore.BLUE + "\n--- Iniciando o processo de atualização de Equipamentos (URL de Monitoramento) via Selenium ---" + Style.RESET_ALL)

    # Carregar planilha
    workbook = load_workbook(caminho_planilha)
    if nome_aba not in workbook.sheetnames:
        raise ValueError(f"Aba '{nome_aba}' não encontrada na planilha. Verifique o nome da aba.")
    sheet = workbook[nome_aba]
    print(Fore.GREEN + f"Aba '{nome_aba}' carregada com sucesso." + Style.RESET_ALL)

    # Obter índices das colunas
    header = [cell.value for cell in sheet[1]]
    try:
        idx_coluna_instalacao = header.index(coluna_instalacao) + 1
        idx_coluna_equipamento = header.index(coluna_equipamento) + 1
    except ValueError as e:
        raise KeyError(f"Uma das colunas especificadas ({coluna_instalacao} ou {coluna_equipamento}) não foi encontrada no cabeçalho: {e}")

    # Configurar WebDriver
    service = webdriver.chrome.service.Service(caminho_chromedriver)
    options = webdriver.ChromeOptions()
    # options.add_argument("--headless") # Descomente para rodar sem interface gráfica
    options.add_argument("--log-level=3")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-extensions")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)

    driver = webdriver.Chrome(service=service, options=options)
    driver.set_page_load_timeout(timeout_selenium_geral)
    print(Fore.GREEN + "WebDriver configurado." + Style.RESET_ALL)

    # Realiza o login inicial
    if not realizar_login(driver, CUSTOMER_VALUE, USERNAME_VALUE, PASSWORD_VALUE, base_url_vectora, timeout_selenium_geral):
        driver.quit()
        exit() # Encerra o script se o login inicial falhar

    print(Fore.BLUE + "\n--- Processando PCs/DMCs (Nomes de Instalação) e buscando URLs de monitoramento na aba Scorpion ---" + Style.RESET_ALL)

    is_on_scorpion_tab = False # Variável para controlar se já estamos na aba Scorpion

    # --- MODIFICAÇÃO CHAVE AQUI: Inclui DMCs e PCs, incluindo "PC Boo", e exclui VRPs genéricas. ---
    PREFIXOS_VALIDOS_INSTALACAO = ["PC", "PC VRP", "PC DMC", "PC Boo", "DMC"]

    # Loop principal para processar cada linha do Excel
    for row_idx in range(2, sheet.max_row + 1):
        celula_instalacao = sheet.cell(row=row_idx, column=idx_coluna_instalacao)
        celula_equipamento = sheet.cell(row=row_idx, column=idx_coluna_equipamento)

        nome_da_instalacao_original = str(celula_instalacao.value).strip() if celula_instalacao.value else ""
        numero_fixo_do_pc_title = str(celula_equipamento.value).strip() if celula_equipamento.value else ""

        url_encontrada_para_atualizar = None
        codigo_equipamento_do_link = None

        # Remove qualquer número (incluindo pontos) seguido por espaços no início da string
        nome_da_instalacao_limpo = re.sub(r'^\s*[\d\.]+\s*', '', nome_da_instalacao_original).strip()

        processar_linha = False
        # Converter para maiúsculas para comparação case-insensitive
        nome_da_instalacao_limpo_upper = nome_da_instalacao_limpo.upper()

        # Adiciona uma regra explícita para IGNORAR "VRP" que não seja "PC VRP"
        # Verifica se o prefixo "VRP" está presente e se o nome NÃO COMEÇA com "PC VRP" após a limpeza.
        if "VRP" in nome_da_instalacao_limpo_upper and not nome_da_instalacao_limpo_upper.startswith("PC VRP"):
            print(Fore.YELLOW + f"     Linha {row_idx}: Ignorando '{nome_da_instalacao_original}' (Contém 'VRP' mas não é 'PC VRP')." + Style.RESET_ALL)
            continue # Pula esta linha

        for prefixo in PREFIXOS_VALIDOS_INSTALACAO:
            # Continua verificando se o prefixo está CONTIDO na string limpa (para casos como "JO PC VRP Godofredo")
            if prefixo.upper() in nome_da_instalacao_limpo_upper:
                processar_linha = True
                break
        
        if not processar_linha:
            print(Fore.YELLOW + f"     Linha {row_idx}: Ignorando '{nome_da_instalacao_original}' (Não corresponde a nenhum prefixo válido após limpeza inicial: {', '.join(PREFIXOS_VALIDOS_INSTALACAO)})." + Style.RESET_ALL)
            continue

        if not numero_fixo_do_pc_title:
            print(Fore.YELLOW + f"     Linha {row_idx}: Ignorando (sem número fixo/texto para encontrar o link)." + Style.RESET_ALL)
            continue

        print(Fore.CYAN + f"\n     Processando Linha {row_idx}: Buscando '{nome_da_instalacao_original}' (Nome Limpo para busca: '{nome_da_instalacao_limpo}', N° Fixo/Texto do Link: {numero_fixo_do_pc_title})" + Style.RESET_ALL)

        try:
            # 1. VERIFICAR SESSÃO E RELOGAR SE NECESSÁRIO
            verificar_e_relogar(driver, CUSTOMER_VALUE, USERNAME_VALUE, PASSWORD_VALUE, base_url_vectora, timeout_selenium_geral)

            # 2. NAVEGAR PARA A ABA SCORPION APENAS SE AINDA NÃO ESTIVER LÁ
            try:
                if not is_on_scorpion_tab:
                    print(Fore.CYAN + f"     Clicando na aba 'Scorpion 35PF'." + Style.RESET_ALL)
                    WebDriverWait(driver, timeout_pequena_espera).until(EC.element_to_be_clickable(SELETOR_ABA_SCORPION)).click()
                    is_on_scorpion_tab = True
                else:
                    print(Fore.CYAN + f"     Já estava na aba 'Scorpion 35PF'. Verificando se a página está pronta." + Style.RESET_ALL)

                WebDriverWait(driver, timeout_selenium_geral).until(EC.presence_of_element_located(SELETOR_SEARCH_BOX_SCORPICO_ESPECIFICO))
                print(Fore.CYAN + "     Caixa de busca da aba Scorpion confirmada." + Style.RESET_ALL)

            except TimeoutException as e:
                print(Fore.RED + f"     ❌ ERRO: Tempo limite excedido ao tentar acessar ou confirmar a aba 'Scorpion 35PF' ou o seu campo de busca: {e.__class__.__name__}: {e}" + Style.RESET_ALL)
                celula_equipamento.value = f"NÃO ATUALIZADO - Erro Aba Scorpion/Busca"
                celula_equipamento.font = Font(color='FF0000', bold=True)
                continue

            # 3. INTERAGIR COM A CAIXA DE BUSCA
            try:
                seletor_busca_atual = SELETOR_SEARCH_BOX_SCORPICO_ESPECIFICO
                WebDriverWait(driver, timeout_pequena_espera).until(EC.element_to_be_clickable(seletor_busca_atual))
                search_box = driver.find_element(*seletor_busca_atual)
                print(Fore.CYAN + "     Caixa de busca detectada e clicável." + Style.RESET_ALL)

                driver.execute_script("arguments[0].value = '';", search_box)
                search_box.send_keys(Keys.CONTROL + "a")
                search_box.send_keys(Keys.DELETE)
                time.sleep(1)

                search_box.send_keys(nome_da_instalacao_limpo)
                print(Fore.CYAN + f"     Digitado '{nome_da_instalacao_limpo}' na caixa de busca. Pressionando ENTER..." + Style.RESET_ALL)
                search_box.send_keys(Keys.ENTER)

                time.sleep(2)

                WebDriverWait(driver, timeout_selenium_geral).until(
                    EC.visibility_of_element_located(SELETOR_TABELA_RESULTS)
                )
                print(Fore.GREEN + "     ✔️ Div wrapper da tabela de resultados visível." + Style.RESET_ALL)

                print(Fore.MAGENTA + f"DEBUG: Realizando captura de tela e page_source para '{nome_da_instalacao_original}' (Texto do Link: '{numero_fixo_do_pc_title}')" + Style.RESET_ALL)
                screenshot_path = f"debug_screenshot_row_{row_idx}_antes_link.png"
                driver.save_screenshot(screenshot_path)
                print(Fore.MAGENTA + f"DEBUG: Captura de tela salva em {screenshot_path}" + Style.RESET_ALL)
                page_source_debug = driver.page_source
                print(Fore.MAGENTA + "DEBUG: Início do Page Source (apenas 5000 primeiros caracteres):\n" + page_source_debug[:5000] + "\n..." + Style.RESET_ALL)

                print(Fore.CYAN + f"     Tentando localizar o link com o TEXTO VISÍVEL '{numero_fixo_do_pc_title}' nos resultados e extrair 'href'..." + Style.RESET_ALL)

            except TimeoutException as e:
                print(Fore.RED + f"     ❌ ERRO: Tempo limite excedido ao interagir com a caixa de busca ou aguardar resultados da div wrapper da tabela: {e.__class__.__name__}: {e}" + Style.RESET_ALL)
                celula_equipamento.value = f"NÃO ATUALIZADO - Erro Busca Scorpion/Resultados"
                celula_equipamento.font = Font(color='FF0000', bold=True)
                continue
            except Exception as e:
                print(Fore.RED + f"     ❌ ERRO inesperado ao interagir com a caixa de busca: {e.__class__.__name__}: {e}" + Style.RESET_ALL)
                celula_equipamento.value = f"NÃO ATUALIZADO - Erro Busca Scorpion Geral"
                celula_equipamento.font = Font(color='FF0000', bold=True)
                continue

            # 4. ENCONTRAR O LINK PELO TEXTO VISÍVEL E EXTRAIR 'href'
            try:
                link_elemento = WebDriverWait(driver, timeout_selenium_geral).until(
                    EC.element_to_be_clickable(get_seletor_link_do_numero_fixo(numero_fixo_do_pc_title))
                )

                codigo_equipamento_do_link = link_elemento.text
                print(Fore.GREEN + f"     ✔️ Texto visível do link (Código do Equipamento) encontrado: '{codigo_equipamento_do_link}'" + Style.RESET_ALL)

                url_do_monitoramento = link_elemento.get_attribute("href")
                print(Fore.GREEN + f"     ✔️ Atributo 'href' (URL de monitoramento) encontrado: '{url_do_monitoramento}'" + Style.RESET_ALL)

                url_encontrada_para_atualizar = url_do_monitoramento

            except (TimeoutException, NoSuchElementException):
                print(Fore.RED + f"     ❌ ERRO: Link com o TEXTO VISÍVEL '{numero_fixo_do_pc_title}' NÃO encontrado nos resultados da busca dentro do tempo esperado." + Style.RESET_ALL)
                celula_equipamento.font = Font(color='FF0000', bold=True)
                celula_equipamento.value = f"NÃO ATUALIZADO - Link '{numero_fixo_do_pc_title}' não encontrado"
                continue
            except Exception as e:
                print(Fore.RED + f"     ❌ ERRO inesperado ao extrair texto ou 'href' do link: {e.__class__.__name__}: {e}" + Style.RESET_ALL)
                celula_equipamento.font = Font(color='FF0000', bold=True)
                celula_equipamento.value = f"NÃO ATUALIZADO - Erro Extração Link"
                continue

        except Exception as e:
            print(Fore.RED + f"     ERRO geral de Selenium ao interagir com a aba Scorpion para '{nome_da_instalacao_original}' (N° Fixo/Texto: {numero_fixo_do_pc_title}): {e.__class__.__name__}: {e}" + Style.RESET_ALL)
            print(Fore.YELLOW + f"     Esta linha não será processada corretamente devido ao erro." + Style.RESET_ALL)
            celula_equipamento.font = Font(color='FF0000', bold=True)
            celula_equipamento.value = f"NÃO ATUALIZADO - Erro Selenium Geral"


        # 6. ATUALIZAR A CÉLULA DO EXCEL com o NÚMERO (do link) e o LINK (Hiperlink)
        if url_encontrada_para_atualizar and codigo_equipamento_do_link:
            celula_equipamento.value = codigo_equipamento_do_link
            celula_equipamento.hyperlink = url_encontrada_para_atualizar
            celula_equipamento.font = Font(color=BLUE, underline='single')
            print(Fore.GREEN + f"     Célula de Equipamento {celula_equipamento.coordinate} atualizada para o número '{celula_equipamento.value}' com hiperlink para '{url_encontrada_para_atualizar}'." + Style.RESET_ALL)
        else:
            if not celula_equipamento.font or celula_equipamento.font.color.rgb != 'FFFF0000':
                print(Fore.RED + f"     ❌ Não foi possível obter URL ou código do equipamento para '{numero_fixo_do_pc_title}' (linha {row_idx}). Mantendo valor original." + Style.RESET_ALL)
                celula_equipamento.value = f"NÃO ATUALIZADO - Processamento Incompleto"
                celula_equipamento.font = Font(color='FF0000', bold=True)


    print(Fore.BLUE + f"\n--- Salvando Planilha Atualizada: {os.path.basename(caminho_planilha)} ---" + Style.RESET_ALL)
    workbook.save(caminho_planilha)
    print(Fore.GREEN + f"Planilha salva com sucesso em '{caminho_planilha}'." + Style.RESET_ALL)

# --- TRATAMENTO DE ERROS FINAIS (APÓS O LOOP PRINCIPAL) ---
except FileNotFoundError as e:
    print(Fore.RED + f"\nERRO: Arquivo não encontrado. Verifique o caminho configurado: {e}" + Style.RESET_ALL)
except KeyError as e:
    print(Fore.RED + f"\nERRO: Coluna não encontrada ou nome incorreta: {e}" + Style.RESET_ALL)
except ValueError as e:
    print(Fore.RED + f"\nERRO: Problema com o nome da aba ou tipo de dado: {e}" + Style.RESET_ALL)
except Exception as e:
    print(Fore.RED + f"\nOcorreu um erro inesperado e não tratado (fora do loop de processamento): {e.__class__.__name__}: {e}" + Style.RESET_ALL)
finally:
    if 'driver' in locals() and driver:
        driver.quit()
        print(Fore.MAGENTA + "\nWebDriver fechado." + Style.RESET_ALL)